
import customtkinter as ctk
import style_register as style
import cv2
from PIL import Image, ImageTk
import os
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import face_recognition
import pickle
import numpy as np


class face_recog_window(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.geometry("1024x600")
        self.title("Face Recognition")
        self.attributes('-topmost', True)
        self.attributes('-fullscreen', True)

        
        self.back_btn = style.btn_style(self, 'Back', self.back_func_btn)
        self.back_btn.place(x=855.7, y=23.5)

        self.arrival_btn_AM = style.btn_style(self, 'Arrival AM', self.arrival_func_AM)
        self.arrival_btn_AM.place(x=629.1, y=300)

        self.departure_btn_AM = style.btn_style(self, 'Departure AM', self.departure_func_AM)
        self.departure_btn_AM.place(x=782.3, y=300)

        self.arrival_btn_PM = style.btn_style(self, 'Arrival PM', self.arrival_func_PM)
        self.arrival_btn_PM.place(x=629.1, y=484.6)

        self.departure_btn_PM = style.btn_style(self, 'Departure PM', self.departure_func_PM)
        self.departure_btn_PM.place(x=782.3, y=484.6)

        self.folderPath = 'test_images'
        self.PathList = os.listdir(self.folderPath)
        self.imgList = []
        self.employee_id = []

        for path in self.PathList:
            self.imgList.append(cv2.imread(os.path.join(self.folderPath, path)))
            
            self.employee_id.append(os.path.splitext(path)[0])
        
        #print(self.employee_id)

    
        def find_encodings(imageList):
            encodeList = []
            for img in imageList:
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                encode = face_recognition.face_encodings(img)[0]
                encodeList.append(encode)

            return encodeList

        #print("Encoding Started...")
        self.encodeListKnown = find_encodings(self.imgList)  
        self.encodeListKnownWithIds = [self.encodeListKnown, self.employee_id]
        #print(self.encodeListKnown)
        #print("Encoding Complete...")

        print("Loading Encode File...")
        self.file_encode = open("EncodeFile.p", 'wb')
        pickle.dump(self.encodeListKnownWithIds, self.file_encode)
        self.file_encode.close()
        #print("File Complete")
        self.file = open('EncodeFile.p', 'rb')
        self.encodeListKnownWithIds = pickle.load(self.file)
        self.encodeListKnown, self.employee_id = self.encodeListKnownWithIds
        self.file.close()
        #print(self.employee_id)
        print("Encode File Loaded")

        self.webcam_label = style.webcam_style_label(self)
        self.webcam_label.place(x=28, y=96, width=474.7, height=475.3)

        self.add_webcam(self.webcam_label)

    def add_webcam(self, label):
        if 'capture' not in self.__dict__:
            self.capture = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()

    def process_webcam(self):
        ret, frame = self.capture.read()

        self.most_recent_capture_arr = frame
        self.img_ = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)

        self.faceCurFrame = face_recognition.face_locations(self.img_)
        self.encodeCurFrame = face_recognition.face_encodings(self.img_, self.faceCurFrame) 
        self.most_recent_capture_pil = Image.fromarray(self.img_)
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        self._label.imgtk = imgtk
        self._label.configure(image=imgtk)
        
        self._label.after(20, self.process_webcam)



    def arrival_func_AM(self):

        for encodeFace, faceLoc in zip(self.encodeCurFrame, self.faceCurFrame):
            self.matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            self.faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            self.matchIndex = np.argmin(self.faceDis)
           
            if self.matches[self.matchIndex]:
                employee_id = self.employee_id[self.matchIndex]
                self.attributes('-topmost', False)
                result = messagebox.askquestion("Match Found", f"Employee ID: {employee_id}\nDo you want your time in now?")
                self.attributes('-topmost', True)
                if result == 'yes':
                    
                    workbook = openpyxl.load_workbook("Data.xlsx")
                    worksheet = workbook.get_sheet_by_name("Attendance")

                    row = worksheet.max_row + 1
                    now = datetime.now()
                    
                            
                    worksheet.cell(row=row, column=1, value=employee_id)
                    worksheet.cell(row=row, column=3, value=now.strftime("%m-%d-%Y"))
                    worksheet.cell(row=row, column=4, value=now.strftime("%A"))
                    worksheet.cell(row=row, column=5, value=now.strftime("%H:%M:%S"))

                     
                    workbook.save('Data.xlsx')  
     
            else:
                self.attributes('-topmost', False)
                result = messagebox.showerror("No Face Match found")
                self.attributes('-topmost', True)
       
        print("Arrival AM just clicked")


    def arrival_func_PM(self):          
       
        for encodeFace, faceLoc in zip(self.encodeCurFrame, self.faceCurFrame):
            self.matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            self.faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            self.matchIndex = np.argmin(self.faceDis)
           
            if self.matches[self.matchIndex]:
                employee_id = self.employee_id[self.matchIndex]
                self.attributes('-topmost', False)
                result = messagebox.askquestion("Match Found", f"Employee ID: {employee_id}\nDo you want to save your time in now?")
                self.attributes('-topmost', True)
                if result == 'yes':
                    
                    workbook = openpyxl.load_workbook("Data.xlsx")
                    worksheet = workbook.get_sheet_by_name("Attendance")
                    now = datetime.now()
                    row = worksheet.max_row
                    worksheet.cell(row=row, column=1, value=employee_id)
                    worksheet.cell(row=row, column=3, value=now.strftime("%m-%d-%Y"))
                    worksheet.cell(row=row, column=4, value=now.strftime("%A"))

                    worksheet.cell(row=row, column=7, value=now.strftime("%H:%M:%S"))

                    workbook.save('Data.xlsx')  
            else:
                self.attributes('-topmost', False)
                result = messagebox.showerror("No Face Match found")
                self.attributes('-topmost', True)
        print("Arrival PM just clicked")


    def departure_func_AM(self):
        for encodeFace, faceLoc in zip(self.encodeCurFrame, self.faceCurFrame):
            self.matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            self.faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            self.matchIndex = np.argmin(self.faceDis)
           
            if self.matches[self.matchIndex]:
                employee_id = self.employee_id[self.matchIndex]
                self.attributes('-topmost', False)
                result = messagebox.askquestion("Match Found", f"Employee ID: {employee_id}\nDo you want to save your time in now?")
                self.attributes('-topmost', True)
                if result == 'yes':
                    
                    workbook = openpyxl.load_workbook("Data.xlsx")
                    worksheet = workbook.get_sheet_by_name("Attendance")
                    now = datetime.now()
                    row = worksheet.max_row

                    worksheet.cell(row=row, column=6, value=now.strftime("%H:%M:%S"))

                    workbook.save('Data.xlsx')  
            else:
                self.attributes('-topmost', False)
                result = messagebox.showerror("No Face Match found")
                self.attributes('-topmost', True)
        

        print("Departure AM just clicked")

    def departure_func_PM(self):
        for encodeFace, faceLoc in zip(self.encodeCurFrame, self.faceCurFrame):
            self.matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
            self.faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

            self.matchIndex = np.argmin(self.faceDis)
           
            if self.matches[self.matchIndex]:
                employee_id = self.employee_id[self.matchIndex]
                self.attributes('-topmost', False)
                result = messagebox.askquestion("Match Found", f"Employee ID: {employee_id}\nDo you want to save your time in now?")
                self.attributes('-topmost', True)
                if result == 'yes':
                    
                    workbook = openpyxl.load_workbook("Data.xlsx")
                    worksheet = workbook.get_sheet_by_name("Attendance")
                    now = datetime.now()
                    row = worksheet.max_row
                    worksheet.cell(row=row, column=8, value=now.strftime("%H:%M:%S"))

                    workbook.save('Data.xlsx')  
            else:
                self.attributes('-topmost', False)
                result = messagebox.showerror("No Face Match found")
                self.attributes('-topmost', True)
        

        print("Departure PM just clicked")


    def back_func_btn(self):
        self.capture.release()
        self.destroy()
        print("Back button just clicked")





