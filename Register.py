import customtkinter as ctk
import style_register as style
import cv2
from PIL import Image, ImageTk
import os
import tkinter
from tkinter import messagebox
import openpyxl
import Face_Recognize

class register_window(ctk.CTkToplevel):
    
    def __init__(self, master):
        super().__init__(master)
        self.geometry("1024x600")
        self.title("Register")
        self.attributes('-topmost', True)

        self.register_label_heading = ctk.CTkLabel(self, text="Register", font=('Poppins', 24), text_color="white")
        self.register_label_heading.place(x=432.5, y=21.5)

        self.back_btn = style.btn_style(self, 'Back', self.back_func_btn)
        self.back_btn.place(x=855.7, y=23.5)

        self.frame = ctk.CTkFrame(self, width=306, height=480, corner_radius=15)
        self.frame.place(x=690, y=98)

        self.employee_pro_label = ctk.CTkLabel(self, text="Employee Profile", font=('Poppins', 24), text_color="white", bg_color="#2B2B2B")
        self.employee_pro_label.place(x=755, y=108)

        self.name_label = ctk.CTkLabel(self, text="Name", font=('Poppins', 12), text_color="white", bg_color="#2B2B2B")
        self.name_label.place(x=705, y=163)

        self.user_label = ctk.CTkLabel(self, text="User ID", font=('Poppins', 12), text_color="white", bg_color="#2B2B2B")
        self.user_label.place(x=705, y=242)

        self.position_label = ctk.CTkLabel(self, text="Position", font=('Poppins', 12), text_color="white", bg_color="#2B2B2B")
        self.position_label.place(x=705, y=321)

        self.status_label = ctk.CTkLabel(self, text="Status", font=('Poppins', 12), text_color="white", bg_color="#2B2B2B")
        self.status_label.place(x=705, y=400)

        self.name_entry = style.entry_style(self, "Enter your name")
        self.name_entry.place(x=705, y=185.5)

        self.user_entry = style.entry_style(self, "Enter your user id")
        self.user_entry.place(x=705, y=264.4)

        self.position_entry = style.entry_style(self, "Enter your position")
        self.position_entry.place(x=705, y=343.3)

        self.status_entry = style.entry_style(self, "Enter your status")
        self.status_entry.place(x=705, y=422.2)

        self.capture_btn = style.frame_btn_style(self, "Capture", self.capture_func_btn,)
        self.capture_btn.place(x=772.3, y=501.3)


        self.webcam_label = style.webcam_style_label(self)
        self.webcam_label.place(x=28, y=96, width=640, height=480)

        self.name = ""
        self.user_id = ""
        self.position = ""
        self.status = ""
        self.capture_image_path = None
        self.btn_confirm = False

        self.add_webcam(self.webcam_label)

    def add_webcam(self, label):
        if 'capture' not in self.__dict__:
            self.capture = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()

    def process_webcam(self):
        ret, frame = self.capture.read()
        self.most_recent_capture_arr = frame
        img_ = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)
        self.most_recent_capture_pil = Image.fromarray(img_)
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        self._label.imgtk = imgtk
        self._label.configure(image=imgtk)
        self._label.after(20, self.process_webcam)
    

    def capture_func_btn(self):
        
        user_id_value = self.user_entry.get() 

        ret, frame = self.capture.read()
        if ret:
            if not os.path.exists('test_images'):
                os.makedirs('test_images')
            path_image = 'test_images/{}.jpg'.format(user_id_value) 
            cv2.imwrite(path_image, frame)

            workbook = openpyxl.load_workbook("Data.xlsx")
            worksheet = workbook.active
            for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
                if row[0].value == user_id_value:
                    self.attributes('-topmost', False)
                    messagebox.showerror("Duplicate ID", "The ID entered already exists")
                    self.attributes('-topmost', True)
                    return
                
            #Create a new window for registration of the images
            self.photo_register = ctk.CTkToplevel()
            self.photo_register.geometry("1024x600")
            self.photo_register.title("Photo Registration")
            self.photo_register.attributes('-topmost', True)
            self.attributes('-topmost', False)
            
            self.image_capture_label = ctk.CTkLabel(self.photo_register, text="")
            self.image_capture_label.place(x=360, y=90)
            
            self.capture_image_path = path_image  
            self.captured_image = Image.open(path_image)
            self.captured_image = ImageTk.PhotoImage(self.captured_image)
            self.image_capture_label.configure(image=self.captured_image)
            self.image_capture_label.image = self.captured_image 

            self.photo_register_label = ctk.CTkLabel(self.photo_register, text="Photo Registration", font=('Poppins', 30), bg_color="#242424")
            self.photo_register_label.place(x=28, y=21.5)

            self.frame_photo = ctk.CTkFrame(self.photo_register, width=306, height=480, corner_radius=15)
            self.frame_photo.place(x=28, y=90.8)

            self.question_label = ctk.CTkLabel(self.photo_register, text="Is this photo okay to you?", font=('Poppins', 24), bg_color="#2B2B2B")
            self.question_label.place(x=50, y=114.8)

            self.yes1_label = ctk.CTkLabel(self.photo_register, text="Click Yes if you are okay with", font=('Poppins', 20), bg_color="#2B2B2B")
            self.yes1_label.place(x=44.7, y=213)
            self.yes2_label = ctk.CTkLabel(self.photo_register, text="this photo and register it", font=('Poppins', 20), bg_color="#2B2B2B")
            self.yes2_label.place(x=68.7, y=238)

            self.no1_label = ctk.CTkLabel(self.photo_register, text="Click No if you want to take a", font=('Poppins', 20), bg_color="#2B2B2B")
            self.no1_label.place(x=44.7, y=300)
            self.no2_label = ctk.CTkLabel(self.photo_register, text="photo again", font=('Poppins', 20), bg_color="#2B2B2B")
            self.no2_label.place(x=125.4, y=326)

            
            self.yes_btn = style.frame_btn_style(self.photo_register, "Yes", self.yes_func_btn)
            self.yes_btn.place(x=118, y=388.4)

            self.no_btn = style.frame_btn_style(self.photo_register, "No", self.no_func_btn)
            self.no_btn.place(x=118, y=466.7)
            
        print("Capture function button just clicked")

        
    def yes_func_btn(self):
        user_id_value = self.user_entry.get()
        
                
        self.name = self.name_entry.get()
        self.user_id = self.user_entry.get()
        self.position = self.position_entry.get()
        self.status = self.status_entry.get()

        workbook = openpyxl.load_workbook("Data.xlsx")
        worksheet = workbook.active
        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=self.name)
        worksheet.cell(row=row, column=2, value=self.user_id)
        worksheet.cell(row=row, column=3, value=self.position)
        worksheet.cell(row=row, column=4, value=self.status)
        workbook.save("Data.xlsx")

        
        print("You are register")
        self.photo_register.destroy()

    def no_func_btn(self):
        self.btn_confirm = False
        os.remove(self.capture_image_path)
        self.reset_capture()
        print("No button just click")
        
    def reset_capture(self):
        self.image_capture_label.configure(image=None)
        self.image_capture_label.image = None
        self.photo_register.destroy()

    def face_func_btn(self):
        self.capture.release()
        Face_Recognize.face_recog_window(ctk.CTkToplevel)
        self.destroy()
        print("Face Recognize button just clicked")

    def about_func_btn(self):
        print("About us button just clicked")
    
    def back_func_btn(self):
        self.capture.release()
        self.destroy()
        print("Back button just clicked")

if __name__ == "__main__":
    root = ctk.CTk()
    app = register_window(root)
    app.mainloop()